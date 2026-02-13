import os
from openpyxl import load_workbook

# 文件夹路径
folder_path = r''  # 请替换为你的文件夹路径

# 遍历文件夹中的所有文件
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        print(f"处理文件: {filename}")
        # 读取 Excel 文件
        file_path = os.path.join(folder_path, filename)
        wb = load_workbook(file_path)
        ws = wb.active

        # 寻找矩阵起始位置
        start_row = None
        start_col = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value == 200:
                    if ws.cell(row=row + 1, column=col - 1).value == 250:
                        start_row = row + 1
                        start_col = col
                        break
            if start_row is not None:
                break

        if start_row is None or start_col is None:
            print(f"未找到矩阵起始位置: {filename}")
            continue

        # 提取矩阵数据
        matrix_data = []
        for i in range(start_row, start_row + 61):
            row_data = []
            for j in range(start_col, start_col + 51):
                cell_value = ws.cell(row=i, column=j).value
                row_data.append(cell_value)
            matrix_data.append(row_data)

        # 创建新的工作表
        ws2 = wb.create_sheet("Matrix")

        # 将矩阵数据写入新工作表
        for i, row_data in enumerate(matrix_data, start=1):
            for j, value in enumerate(row_data, start=1):
                ws2.cell(row=i, column=j, value=value)

        # 保存 Excel 文件
        wb.save(file_path)
        print(f"{filename} 处理完毕并保存。")
